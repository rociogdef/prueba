import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
import json
import csv
import sys
from sqlalchemy import create_engine
import re



#wb = load_workbook(filename = 'excel1.xlsx')

try:
   wb = load_workbook(filename = sys.argv[1])
except Exception as ex:
   print("No es posible leer el archivo de entrada")
   exit(1)
    

ws=wb.active


colname=list(ws.rows)[2]  #nombre de las columnas primer renglon 
tiposcol=list(ws.rows)[3]


def valida_dato(dato,valor):
   if dato.lower().strip() == valor:
     return True
   return False

l_colname=['examen técnico desarrollador backend','cliente','# contrato','fecha de compra','ciudad','empresa','listado de adeudos']
l_tipos=['(alfabético)','(alfanumérico)','(fecha)','(alfabético)','(alfanumérico)','(numérico)']
def Valida_encabezado(ws):
   if not valida_dato(ws['B1'].value,l_colname[0]):
      print("Error en titulo1")
      return False
   if not valida_dato(ws['B2'].value,l_colname[6]):
      print("Error en titulo2")
      return False
   for i in range(1,5):
      if not valida_dato(colname[i].value,l_colname[i]): 
         print("Error en nombre de columna "+colname[i].value) 
         return False
   for i in range(1,5):
      if not valida_dato(tiposcol[i].value,l_tipos[i-1]): 
         print("Error en tipo de columna "+tiposcol[i].value) 
         return False         
   return True

#Genera el archivo .csv sin el encabezado
def Transforma_csv(sh): 
   with open('adeudos_aux.csv', 'w', newline="") as f:  
      c = csv.writer(f)    
      #for r in sh.iter_rows(min_row=5, min_col=2, max_col=7,max_row=20):
      for r in sh.iter_rows(min_row=5, min_col=2, max_col=7):
        # print(r)
         c.writerow([cell.value for cell in r])
       
def CargaCSV_tabla():
   with open('aux2.csv','r') as f:
    filas=tuple(csv.reader(f))
   engine = create_engine('postgresql+pg8000://postgres:facil680@localhost:5432/bd_chbous')
   try:
     connection = engine.connect()
   except Exception as ex:
    print(ex)
   for r in filas:  
      sql_insert="insert into challenge_bous.adeudos values "+ str(r).replace('[','(').replace(']',')')
      print(sql_insert)
      try:
         rs = connection.execute(sql_insert)          
      except Exception as ex:
        print(ex)
   print(rs)     
   rs.close()
   

if __name__ == '__main__':
  if not Valida_encabezado(ws):
     exit(1)
  Transforma_csv(ws)   
 # CargaCSV_tabla()